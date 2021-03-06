# 20201230
# choi_jinwook
# jinsrobot@naver.com
# ver - 0.9.4

import os
import openpyxl
from openpyxl.styles import PatternFill


def calRate(this, point):
    result = this / int(people) * 100
    return round(result, point)


def exit():
    if people == 0:
        print("아직 입력된게 아무것도 없오!")
        print("끈다!")
    else:
        show()


def excelOutput():
    wb = openpyxl.Workbook()
    sheet_result = wb['Sheet']
    sheet_result.title = '찡긋 ><'

    print("================")
    print("==excel output==")
    print("================")
    print()
    name = ""
    for character in input("파일 이름을 입력해주세요 >> "):
        if character.isalnum():
            name += character


    deepblueFill = PatternFill(start_color='FF538DD5',
                               end_color='FF538DD5',
                               fill_type='solid')

    blueFill = PatternFill(start_color='FF95B3D7',
                           end_color='FF95B3D7',
                           fill_type='solid')

    lightblueFill = PatternFill(start_color='FFC5D9F1',
                                end_color='FFC5D9F1',
                                fill_type='solid')

    sheet_result.cell(row = 2, column = options + 5).value = "소숫점"
    sheet_result.cell(row = 2, column = options + 5).fill = blueFill
    sheet_result.cell(row = 3, column = options + 5).value = point
    sheet_result.cell(row = 3, column = options + 5).fill = lightblueFill


    for t in range(0, options + 1):
        sheet_result.cell(row=1, column=t + 2).value = "#" + str(t)
        sheet_result.cell(row=1, column=t + 2).fill = deepblueFill

    for t in range(int(questions)):
        sheet_result.cell(row=2 * (t + 1), column=1).value = str(t + 1) + "번"
        sheet_result.cell(row=2 * (t + 1), column=1).fill = deepblueFill
        sheet_result.cell(row=2 * (t + 2) - 1, column=1).fill = deepblueFill

    for x in range(int(questions)):
        sheet_result.cell(row=2 * (x + 1), column=(options + 3)).value = people
        for y in range(0, options + 1):
            if y == 10:
                k = "t"
            else:
                k = str(y)
            sheet_result.cell(row = 2 * (x + 1), column = y + 2).value = array[x][k]
            sheet_result.cell(row = 2 * (x + 2) - 1, column = y + 2).value = ('%s%s%d%s%s%d%s%d%s%s%d%s' % (
                "= ROUND(", excel_alpha[y + 2], (2 * (x + 1)), "/",
                excel_alpha[options + 3], (2 * (x + 1)), "*", 100, ",",
                excel_alpha[options + 5], 3, ")"))         # = round

            if array[x][k] != 0:
                sheet_result.cell(row = 2 * (x + 1), column=y + 2).fill = blueFill
                sheet_result.cell(row = 2 * (x + 2) - 1, column=y + 2).fill = lightblueFill

        sheet_result.cell(row = 2 * (x + 1), column = options + 3).value = ('%s%s%d%s%s%d%s' % (
            "= SUM(", excel_alpha[2], (2 * (x + 1)), ":",
            excel_alpha[options + 2], (2 * (x + 1)), ")"))
        sheet_result.cell(row=2 * (x + 2) - 1, column=options + 3).value = ('%s%s%d%s%s%d%s%d%s' % (
            "= ROUND(SUM(", excel_alpha[2], (2 * (x + 2) - 1), ":",
            excel_alpha[options + 2], (2 * (x + 2) - 1), "),",0,")"))

    wb.save(name + ".xlsx")
    print()
    print("================")
    print("==저장되었습니다.==")
    print("================")
    print()


def excel_sum(row1, column1, row2, column2):
    sum_stc = ('%s%s%d%s%s%d%s' % (
            "= SUM(", excel_alpha[row1], column1, ":",
            excel_alpha[row2], column2, ")"))
    return sum_stc


def excel_round(num, point):
    # sum_round = ('%s%d%s%s%d%s%d%s%s%d%s' % (
    #             "= ROUND(", excel_alpha[y + 2], (2 * (x + 1)), "/",
    #             excel_alpha[options + 3],(2 * (x + 1)), "*", 100, ",",
    #             excel_alpha[options + 6],2, ")"))         # = round
    return sum_round


def show():
    print("------@ 결과 @------")
    print("참여 인원 -> " + str(people) + "명")
    print('%s%d%s' % ("무효표 -> ", abnormal, "개"))
    for x in range(int(questions)):
        print('%d%s' % (x + 1, "."), end=" ")

        for y in range(0, options + 1):
            if y == 10:
                k = "t"
            else:
                k = str(y)
            print('%2d%-2s%4d%-2s' % (y, ":", array[x][k], "ea"), end=" ")
        print()
        print("%3s" % "", end="")
        for y in range(0, options + 1):
            if y == 10:
                k = "t"
            else:
                k = str(y)
            print('%2s%6.2f%-2s' % ("(", calRate(array[x][k], point), "%)"), end=" ")
        print()
    return


def snap():
    print("=================================================")
    print("VVsncVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVVV")
    print("=================================================")
    print("+ 사용법")
    print("+ 총 문제 갯수와 선택 가능한 보기 갯수의 최댓값 입력")
    print("+ 문제수는 제한이 없지만 보기는 0 ~ 10까지 가능")
    print("+ 선택한 숫자를 차례로 입력")
    print("+ 선택지가 10일 경우 10 대신 't' 입력")
    print("+ ex) 12t34 + enter")
    print("+ 보기 항목에 없는 값이나 문제보다 더 많은 값 입력시 재입력")
    print("+ 무효표가 있을 경울 'x'를 입력하면 무효표 추가 가능")
    print("+ 사용법이나 도움이 필요할 경우 'help' 입력")
    print("+ 's'입력시 excel파일로 저장")
    print("+ 다시 입력해서 저장하는경우 기존 파일은 닫고 실행해주세요!")
    print("+ 'm'입력시 중간점검 가능하고 'exit'입력시 종료")
    print("=================================================")
    print("AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA")
    print("==========================================v0.9.4=")
    print()


def snap_help():
    print("------@ 도움말 @------")
    print()
    print("    m -> 중간점검")
    print("    x -> 무효표 입력")
    print("    s -> excel로 저장")
    print("    exit -> 종료")
    print()


if __name__ == '__main__':
    array = []
    ques_dic = {}
    point = 2  # 소숫점
    excel_alpha = ["null", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q"]

    snap()

    while True:
        questions = input("설문지에 몇가지 문제가 존재하나요?? \n => ")
        options = input("보기 갯수는 몇개인가요? (최대값을 입력해주세요!) \n => ")

        if not questions.isdigit() or not options.isdigit():
            print("++++++++++++++++++++++++++++++++++++")
            print("잘못된 입력입니다. 다시 입력해주세요.")
            print("++++++++++++++++++++++++++++++++++++")
            print()
        elif int(options) > 10 or int(options) <= 0:
            print("++++++++++++++++++++++++++++++++++++++++++++++++")
            print("보기 갯수는 최대 10까지 가능합니다. 확인해주세요.")
            print("++++++++++++++++++++++++++++++++++++++++++++++++")
            print()
        else:
            break

    people = 0
    abnormal = 0
    questions = int(questions)
    options = int(options)

    for x in range(0, options + 1):
        if x == 10:
            x = "t"
        else:
            pass
        ques_dic[str(x)] = 0

    for x in range(0, questions):
        array.append(ques_dic.copy())

    right = True

    print()
    while True:

        if not right:
            people - 1
            right = True

        print('%s %d%s' % ("======", people + 1, "번째 입력"))
        print("**", end="")

        for k in range(0, questions):
            print("*", end="")
        print()

        result = input(">>")

        if result == "x":
            abnormal = abnormal + 1
            print('%s%d%s' % ("무효 설문지가 ", abnormal, "개 입력되었습니다."))
            continue

        elif result == "s":
            excelOutput()

        elif result == "m":
            right = False
            print()
            if people == 0:
                print("아직 입력된게 아무것도 없오!")
            else:
                show()
            print()

        elif result == "help":
            right = False
            print()
            snap_help()
            print()

        elif result == "exit":
            right = False
            print()
            is_exit = input("정말 종료하시겠습니까? (y/n) >> ")
            if is_exit == "Y" or is_exit == "y":
                print("감사합니다.")
                exit()
                break
            elif is_exit == "N" or is_exit == "n":
                print("입력을 계속해서 진행합니다.")
            else:
                print("잘못된 입력입니다.")

            print()

        elif len(result) != questions:
            right = False
            print("덜이나 더 입력됐어용 > ")

        else:
            for i in range(len(result)):
                right = False
                for j in ques_dic.keys():
                    if str(result[i]) == j:
                        right = True

                if not right:
                    print("잘못된 입력입니다.")
                    break

            if right:
                for y in range(0, questions):
                    array[y][str(result[y])] = array[y][str(result[y])] + 1
                people = people + 1

            else:
                pass

    os.system("pause")